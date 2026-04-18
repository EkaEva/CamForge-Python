"""CamForge 导出管理器

处理 TIFF、SVG、CSV、Excel、GIF 等格式的导出逻辑。
"""

import os
import csv as csv_mod
import numpy as np
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.figure import Figure
from matplotlib.gridspec import GridSpec
import matplotlib.pyplot as plt

try:
    import openpyxl
except ImportError:
    openpyxl = None

from i18n import t
from ui.constants import STATIC_DPI, MAX_PRESSURE_ANGLE
from ui.animation import generate_gif_frames


class ExportManager:
    """处理图像/数据导出（TIFF, SVG, CSV, Excel, GIF）。"""

    def __init__(self, root, sim_data_lock):
        self.root = root
        self._sim_data_lock = sim_data_lock

    def download(self, sim_data, toggles, draw_funcs, lang, tk_font_family,
                 axes_ref, status_var, folder):
        """下载勾选的图片。

        Parameters
        ----------
        sim_data : dict
            仿真数据。
        toggles : dict
            导出勾选项 BooleanVar 字典，键: s, v, a, profile, anim, excel,
            alpha, curvature, svg, csv。
        draw_funcs : dict
            绘图函数字典，键: displacement, velocity, acceleration,
            profile, pressure_angle, curvature_radius。
        lang : str
            当前语言代码。
        tk_font_family : str
            tkinter 字体族名。
        axes_ref : dict
            动画坐标轴引用，键: ax_anim。
        status_var : tk.StringVar
            状态栏变量。
        folder : str
            导出目标文件夹路径。
        """
        if not any(toggles.values()):
            status_var.set(t("status.no_download_selection", lang))
            return

        if sim_data is None:
            status_var.set(t("status.run_first", lang))
            return

        dpi = STATIC_DPI
        saved = []
        data = sim_data
        errors = []

        if toggles.get('s'):
            try:
                fig_s = Figure(figsize=(6, 4), dpi=dpi)
                ax_s = fig_s.add_subplot(111)
                draw_funcs['displacement'](ax_s, data)
                filename_s = t("export.filename.displacement", lang) + ".tiff"
                fig_s.savefig(os.path.join(folder, filename_s), dpi=dpi, bbox_inches='tight', format='tiff')
                plt.close(fig_s)
                saved.append(filename_s)
            except Exception as exc:
                errors.append(f"displacement: {exc}")

        if toggles.get('v'):
            try:
                fig_v = Figure(figsize=(6, 4), dpi=dpi)
                ax_v = fig_v.add_subplot(111)
                draw_funcs['velocity'](ax_v, data)
                filename_v = t("export.filename.velocity", lang) + ".tiff"
                fig_v.savefig(os.path.join(folder, filename_v), dpi=dpi, bbox_inches='tight', format='tiff')
                plt.close(fig_v)
                saved.append(filename_v)
            except Exception as exc:
                errors.append(f"velocity: {exc}")

        if toggles.get('a'):
            try:
                fig_a = Figure(figsize=(6, 4), dpi=dpi)
                ax_a = fig_a.add_subplot(111)
                draw_funcs['acceleration'](ax_a, data)
                filename_a = t("export.filename.acceleration", lang) + ".tiff"
                fig_a.savefig(os.path.join(folder, filename_a), dpi=dpi, bbox_inches='tight', format='tiff')
                plt.close(fig_a)
                saved.append(filename_a)
            except Exception as exc:
                errors.append(f"acceleration: {exc}")

        if toggles.get('profile'):
            try:
                fig_p = Figure(figsize=(6, 6), dpi=dpi)
                ax_p = fig_p.add_subplot(111)
                draw_funcs['profile'](ax_p, data)
                filename_p = t("export.filename.profile", lang) + ".tiff"
                fig_p.savefig(os.path.join(folder, filename_p), dpi=dpi, bbox_inches='tight', format='tiff')
                plt.close(fig_p)
                saved.append(filename_p)
            except Exception as exc:
                errors.append(f"profile: {exc}")

        if toggles.get('alpha'):
            try:
                fig_alpha = Figure(figsize=(6, 4), dpi=dpi)
                ax_alpha = fig_alpha.add_subplot(111)
                draw_funcs['pressure_angle'](ax_alpha, data)
                filename_alpha = t("export.filename.pressure_angle", lang) + ".tiff"
                fig_alpha.savefig(os.path.join(folder, filename_alpha), dpi=dpi, bbox_inches='tight', format='tiff')
                plt.close(fig_alpha)
                saved.append(filename_alpha)
            except Exception as exc:
                errors.append(f"pressure_angle: {exc}")

        if toggles.get('curvature'):
            try:
                fig_rho = Figure(figsize=(6, 4), dpi=dpi)
                ax_rho = fig_rho.add_subplot(111)
                draw_funcs['curvature_radius'](ax_rho, data)
                filename_rho = t("export.filename.curvature", lang) + ".tiff"
                fig_rho.savefig(os.path.join(folder, filename_rho), dpi=dpi, bbox_inches='tight', format='tiff')
                plt.close(fig_rho)
                saved.append(filename_rho)
            except Exception as exc:
                errors.append(f"curvature: {exc}")

        # SVG 矢量图导出
        if toggles.get('svg'):
            try:
                fig_svg = Figure(figsize=(14, 7), dpi=100)
                gs_svg = GridSpec(2, 2, figure=fig_svg, wspace=0.30, hspace=0.35)
                ax_s_svg = fig_svg.add_subplot(gs_svg[0, 0])
                ax_v_svg = fig_svg.add_subplot(gs_svg[0, 1])
                ax_a_svg = fig_svg.add_subplot(gs_svg[1, 0])
                ax_p_svg = fig_svg.add_subplot(gs_svg[1, 1])
                draw_funcs['displacement'](ax_s_svg, data)
                draw_funcs['velocity'](ax_v_svg, data)
                draw_funcs['acceleration'](ax_a_svg, data)
                draw_funcs['profile'](ax_p_svg, data)
                filename_svg = "camforge_all.svg"
                fig_svg.savefig(os.path.join(folder, filename_svg), format='svg', bbox_inches='tight')
                plt.close(fig_svg)
                saved.append(filename_svg)
            except Exception as exc:
                errors.append(f"svg: {exc}")

        # CSV 数据表导出
        if toggles.get('csv'):
            try:
                filename_csv = "camforge_data.csv"
                filepath_csv = os.path.join(folder, filename_csv)
                delta_deg = data['delta_deg']
                s_arr = data['s']
                v_arr = data['v']
                a_arr = data['a']
                x_arr = data['x']
                y_arr = data['y']
                R_arr = np.hypot(x_arr, y_arr)
                alpha_arr = data['alpha_all']
                with open(filepath_csv, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv_mod.writer(f)
                    writer.writerow(['delta_deg', 's_mm', 'v_mm_s', 'a_mm_s2',
                                     'x_mm', 'y_mm', 'R_mm', 'alpha_deg'])
                    for i in range(len(delta_deg)):
                        writer.writerow([
                            round(delta_deg[i], 2), round(s_arr[i], 4),
                            round(v_arr[i], 4), round(a_arr[i], 4),
                            round(x_arr[i], 4), round(y_arr[i], 4),
                            round(R_arr[i], 4), round(alpha_arr[i], 4),
                        ])
                saved.append(filename_csv)
            except Exception as exc:
                errors.append(f"csv: {exc}")

        if toggles.get('excel'):
            self._export_excel(folder, saved, data, lang, errors)

        return saved, errors, folder

    def _export_excel(self, folder, saved_list, data, lang, errors):
        """导出凸轮数据为 Excel 表格。"""
        if openpyxl is None:
            errors.append("openpyxl_missing")
            return

        try:
            delta_deg = data['delta_deg']
            v = data['v']
            a = data['a']
            x = data['x']
            y = data['y']
            R = np.hypot(x, y)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = t("excel.sheet_name", lang)

            headers = [
                t("excel.col.delta", lang),
                t("excel.col.radius", lang),
                t("excel.col.velocity", lang),
                t("excel.col.acceleration", lang),
            ]
            ws.append(headers)

            for i in range(len(delta_deg)):
                ws.append([round(delta_deg[i], 1), round(R[i], 4), round(v[i], 4), round(a[i], 4)])

            for col in range(1, 5):
                max_len = len(str(ws.cell(row=1, column=col).value))
                for row in range(2, min(10, len(delta_deg) + 2)):
                    cell_len = len(str(ws.cell(row=row, column=col).value))
                    if cell_len > max_len:
                        max_len = cell_len
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_len + 4

            filename = t("export.filename.excel", lang) + ".xlsx"
            filepath = os.path.join(folder, filename)
            wb.save(filepath)
            saved_list.append(filename)
        except Exception as exc:
            errors.append(f"excel: {exc}")
